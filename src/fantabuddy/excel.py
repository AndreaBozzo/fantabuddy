from __future__ import annotations

import hashlib
import re
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel

EXPECTED_SHEETS = (
    "Tutti",
    "Portieri",
    "Difensori",
    "Centrocampisti",
    "Attaccanti",
    "Ceduti",
)
EXPECTED_HEADERS = (
    "Id",
    "R",
    "RM",
    "Nome",
    "Squadra",
    "Qt.A",
    "Qt.I",
    "Diff.",
    "Qt.A M",
    "Qt.I M",
    "Diff.M",
    "FVM",
    "FVM M",
)
SEASON_PATTERN = re.compile(r"(?:Stagione[_ -])?(20\d{2})[_ -](\d{2,4})", re.IGNORECASE)


class ListoneRecord(BaseModel):
    fantacalcio_id: int
    classic_role: str
    mantra_roles: str
    name: str
    team: str
    quote_current: int
    quote_initial: int
    quote_diff: int
    mantra_quote_current: int
    mantra_quote_initial: int
    mantra_quote_diff: int
    fvm: int
    fvm_mantra: int
    status: str
    source_sheet: str
    source_row: int


class ListoneImport(BaseModel):
    snapshot_id: str
    season: str
    source_path: Path
    source_filename: str
    checksum: str
    source_modified_at: datetime
    records: list[ListoneRecord]

    @property
    def active_count(self) -> int:
        return sum(record.status == "active" for record in self.records)

    @property
    def ceduti_count(self) -> int:
        return sum(record.status == "ceduto" for record in self.records)


def normalize_season(start: int, end: int) -> str:
    end_short = end if end < 100 else end % 100
    if end_short != (start + 1) % 100:
        raise ValueError(f"stagione non consecutiva: {start}/{end}")
    return f"{start}/{end_short:02d}"


def season_start_year(season: str) -> int:
    return int(season.split("/", maxsplit=1)[0])


def infer_season(path: Path, title: object | None = None) -> str:
    for candidate in (path.stem, str(title or "")):
        match = SEASON_PATTERN.search(candidate)
        if match:
            return normalize_season(int(match.group(1)), int(match.group(2)))
    raise ValueError(f"impossibile ricavare la stagione da {path.name}")


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _as_int(value: object, field: str, row_number: int) -> int:
    if value is None or value == "":
        raise ValueError(f"valore mancante per {field} alla riga {row_number}")
    try:
        return int(str(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"valore non intero per {field} alla riga {row_number}: {value!r}"
        ) from exc


def _read_sheet(workbook: object, name: str, status: str) -> list[ListoneRecord]:
    sheet = workbook[name]  # type: ignore[index]
    headers = tuple(cell.value for cell in sheet[2])
    if headers != EXPECTED_HEADERS:
        raise ValueError(
            f"schema inatteso nel foglio {name}: atteso {EXPECTED_HEADERS}, trovato {headers}"
        )

    records: list[ListoneRecord] = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=3, max_col=len(EXPECTED_HEADERS), values_only=True), start=3
    ):
        if not any(value is not None for value in values):
            continue
        record = ListoneRecord(
            fantacalcio_id=_as_int(values[0], "Id", row_number),
            classic_role=str(values[1]).strip(),
            mantra_roles=str(values[2] or "").strip(),
            name=str(values[3]).strip(),
            team=str(values[4]).strip(),
            quote_current=_as_int(values[5], "Qt.A", row_number),
            quote_initial=_as_int(values[6], "Qt.I", row_number),
            quote_diff=_as_int(values[7], "Diff.", row_number),
            mantra_quote_current=_as_int(values[8], "Qt.A M", row_number),
            mantra_quote_initial=_as_int(values[9], "Qt.I M", row_number),
            mantra_quote_diff=_as_int(values[10], "Diff.M", row_number),
            fvm=_as_int(values[11], "FVM", row_number),
            fvm_mantra=_as_int(values[12], "FVM M", row_number),
            status=status,
            source_sheet=name,
            source_row=row_number,
        )
        if record.classic_role not in {"P", "D", "C", "A"}:
            raise ValueError(
                f"ruolo Classic non valido alla riga {row_number}: {record.classic_role}"
            )
        if record.quote_current - record.quote_initial != record.quote_diff:
            raise ValueError(f"Diff. incoerente per ID {record.fantacalcio_id}")
        if record.mantra_quote_current - record.mantra_quote_initial != record.mantra_quote_diff:
            raise ValueError(f"Diff.M incoerente per ID {record.fantacalcio_id}")
        records.append(record)
    return records


def read_listone(path: Path) -> ListoneImport:
    path = path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"formato non supportato: {path.suffix}; serve .xlsx")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if tuple(workbook.sheetnames) != EXPECTED_SHEETS:
            raise ValueError(
                f"fogli inattesi in {path.name}: attesi {EXPECTED_SHEETS}, "
                f"trovati {workbook.sheetnames}"
            )
        title = workbook["Tutti"].cell(row=1, column=1).value
        season = infer_season(path, title)
        records = _read_sheet(workbook, "Tutti", "active")
        records.extend(_read_sheet(workbook, "Ceduti", "ceduto"))
    finally:
        workbook.close()

    ids = [record.fantacalcio_id for record in records]
    duplicated = sorted({player_id for player_id in ids if ids.count(player_id) > 1})
    if duplicated:
        raise ValueError(f"ID duplicati tra Tutti e Ceduti: {duplicated[:20]}")

    checksum = file_sha256(path)
    return ListoneImport(
        snapshot_id=f"listone-{season.replace('/', '-')}-{checksum[:16]}",
        season=season,
        source_path=path,
        source_filename=path.name,
        checksum=checksum,
        source_modified_at=datetime.fromtimestamp(path.stat().st_mtime, tz=UTC),
        records=records,
    )
